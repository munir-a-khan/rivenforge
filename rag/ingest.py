"""
Parse good_rolls.xlsx into a local JSON index + TF-IDF model.

No torch, no onnxruntime, no sentence-transformers, no DLL dependencies.

Two files are written to data/:
  - riven_index.json  : list of weapon dicts (structured data)
  - tfidf_model.json  : vocabulary + IDF weights for similarity search
"""

import json
import math
import os
import re
import sys
from collections import Counter

from core.contracts import WeaponEntryDict

DATA_DIR   = os.path.join(os.path.dirname(__file__), "..", "data")
XLSX_PATH  = os.path.join(DATA_DIR, "good_rolls.xlsx")
ALIAS_PATH = os.path.join(DATA_DIR, "stat_aliases.json")
INDEX_PATH = os.path.join(DATA_DIR, "riven_index.json")
TFIDF_PATH = os.path.join(DATA_DIR, "tfidf_model.json")

WEAPON_SHEETS = ["primary", "secondary", "melee", "archgun", "robotic", "stat sticks"]

with open(ALIAS_PATH) as f:
    ALIASES = json.load(f)

FULL_NAMES = {v.lower(): v for v in ALIASES.values()}


def expand_abbrevs(text: str) -> str:
    if not text:
        return ""
    parts = re.split(r"[/\s,]+", text.strip())
    expanded = []
    for p in parts:
        p = p.strip()
        if not p:
            continue
        expanded.append(ALIASES.get(p.upper(), p))
    return ", ".join(expanded)


def parse_sheet(ws, sheet_name: str) -> list[dict]:
    weapons = []
    for row in ws.iter_rows(values_only=True):
        if not row[0]:
            continue
        cell_a = str(row[0]).strip()
        if cell_a.upper() in ("WEAPON", "NAME", ""):
            continue

        pos_raw = " ".join(filter(None, [
            str(row[1]).strip() if row[1] else "",
            str(row[2]).strip() if len(row) > 2 and row[2] else "",
        ]))
        neg_raw = str(row[3]).strip() if len(row) > 3 and row[3] else ""
        notes   = str(row[4]).strip() if len(row) > 4 and row[4] else ""

        pos_expanded = expand_abbrevs(pos_raw)
        neg_expanded = expand_abbrevs(neg_raw)

        pos_list = [s.strip() for s in pos_expanded.split(",") if s.strip()]
        neg_list = [s.strip() for s in neg_expanded.split(",") if s.strip()]

        weapons.append({
            "weapon":       cell_a,
            "weapon_type":  sheet_name,
            "positives_raw": pos_raw,
            "negatives_raw": neg_raw,
            "positives":    pos_list,
            "negatives":    neg_list,
            "notes":        notes,
            "text_chunk":   _build_chunk(cell_a, sheet_name, pos_list, neg_list, notes, pos_raw, neg_raw),
        })
    return weapons


def _build_chunk(weapon, wtype, pos_list, neg_list, notes, pos_raw, neg_raw) -> str:
    return (
        f"Weapon: {weapon} ({wtype}). "
        f"Desired positives: {', '.join(pos_list) or pos_raw or 'N/A'}. "
        f"Acceptable negatives: {', '.join(neg_list) or 'any'}. "
        f"Notes: {notes or 'none'}."
    )


# ── TF-IDF helpers ────────────────────────────────────────────────────────────

def _tokenize(text: str) -> list[str]:
    return [t for t in re.split(r"[^a-z0-9]+", text.lower()) if t]


def build_tfidf(corpus: list[str]) -> dict:
    """Return {vocab: [str], idf: [float]} for the corpus."""
    N = len(corpus)
    df: Counter = Counter()
    for doc in corpus:
        df.update(set(_tokenize(doc)))
    vocab = sorted(df)
    idf   = [math.log((N + 1) / (df[t] + 1)) + 1 for t in vocab]
    return {"vocab": vocab, "idf": idf}


def save_tfidf(model: dict):
    with open(TFIDF_PATH, "w") as f:
        json.dump(model, f)


# ── Main ingest ───────────────────────────────────────────────────────────────

def ingest(progress_cb=None) -> int:
    """
    Parse the xlsx and write riven_index.json + tfidf_model.json.
    No ML libraries required — pure Python.
    """
    import openpyxl

    if not os.path.exists(XLSX_PATH):
        if not os.path.exists(INDEX_PATH):
            raise FileNotFoundError("good_rolls.xlsx is not available and no riven_index.json exists")

        with open(INDEX_PATH, encoding="utf-8") as f:
            existing_entries = json.load(f)

        corpus = [entry.get("text_chunk", "") for entry in existing_entries]
        save_tfidf(build_tfidf(corpus))
        total = len(existing_entries)
        if progress_cb:
            progress_cb(total, total)
        print(f"[ingest] Rebuilt TF-IDF from existing {INDEX_PATH} ({total} entries)")
        return total

    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

    all_entries: list[dict] = []
    for sheet_name in WEAPON_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        entries = parse_sheet(wb[sheet_name], sheet_name)
        all_entries.extend(entries)

    total = len(all_entries)
    print(f"[ingest] Parsed {total} entries")

    # Write structured index
    with open(INDEX_PATH, "w", encoding="utf-8") as f:
        json.dump(all_entries, f, ensure_ascii=False, indent=1)

    # Build and save TF-IDF model over text_chunk corpus
    corpus = [e["text_chunk"] for e in all_entries]
    model  = build_tfidf(corpus)
    save_tfidf(model)

    if progress_cb:
        progress_cb(total, total)

    print(f"[ingest] Done — {INDEX_PATH} + {TFIDF_PATH}")
    return total


# ── Query helpers (used by rag.py) ────────────────────────────────────────────

_index: list[dict] | None = None
_tfidf_vocab: list[str]   = []
_tfidf_idf:   list[float] = []
_doc_vecs:    list[list[float]] = []


def _load_index():
    global _index, _tfidf_vocab, _tfidf_idf, _doc_vecs
    if _index is not None:
        return

    if not os.path.exists(INDEX_PATH):
        raise FileNotFoundError("riven_index.json not found — run ingest first")

    with open(INDEX_PATH, encoding="utf-8") as f:
        _index = json.load(f)

    if os.path.exists(TFIDF_PATH):
        with open(TFIDF_PATH) as f:
            model = json.load(f)
        _tfidf_vocab = model["vocab"]
        _tfidf_idf   = model["idf"]
        # Pre-compute all doc vectors
        _doc_vecs = [_vectorize(e["text_chunk"]) for e in _index]


def _vectorize(text: str) -> list[float]:
    tokens = _tokenize(text)
    tf = Counter(tokens)
    total = max(sum(tf.values()), 1)
    vec = [(tf[t] / total) * idf for t, idf in zip(_tfidf_vocab, _tfidf_idf, strict=True)]
    norm = math.sqrt(sum(v * v for v in vec)) or 1.0
    return [v / norm for v in vec]


def _cosine(a: list[float], b: list[float]) -> float:
    return sum(x * y for x, y in zip(a, b, strict=False))


def search(query: str, n: int = 3) -> list[WeaponEntryDict]:
    """Return top-n entries most similar to query."""
    _load_index()
    if not _tfidf_vocab:
        return []
    qvec = _vectorize(query)
    scores = [(_cosine(qvec, dv), i) for i, dv in enumerate(_doc_vecs)]
    scores.sort(reverse=True)
    return [_index[i] for _, i in scores[:n]]


def all_weapons() -> list[WeaponEntryDict]:
    """Return all weapon entries (from index or xlsx fallback)."""
    if os.path.exists(INDEX_PATH):
        try:
            _load_index()
            return list(_index)
        except Exception:
            pass
    # Fallback: parse xlsx directly when the source workbook exists locally.
    if not os.path.exists(XLSX_PATH):
        return []

    import openpyxl
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    out = []
    for sheet_name in WEAPON_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        out.extend(parse_sheet(wb[sheet_name], sheet_name))
    return out


def weapon_lookup(weapon_name: str) -> list[WeaponEntryDict]:
    """Return all entries for a specific weapon name (case-insensitive)."""
    return [w for w in all_weapons()
            if w["weapon"].lower() == weapon_name.lower()]


def reset():
    """Force reload on next access."""
    global _index, _doc_vecs
    _index    = None
    _doc_vecs = []


def is_ready() -> bool:
    return os.path.exists(INDEX_PATH) and os.path.exists(TFIDF_PATH)


if __name__ == "__main__":
    sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
    print("[ingest] Building index...")
    n = ingest()
    print(f"[ingest] Done: {n} entries")
